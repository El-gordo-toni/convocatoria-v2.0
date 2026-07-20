from flask import Flask, render_template, request, redirect, session, send_file, jsonify, flash
from datetime import datetime, timedelta
from flask_sqlalchemy import SQLAlchemy
from flask_socketio import SocketIO
from sqlalchemy import event
from sqlalchemy.engine import Engine
from sqlalchemy.exc import IntegrityError, OperationalError
from io import BytesIO
from pathlib import Path
import os
import re
import sqlite3
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "super_secret_key")
ADMIN_PASSWORD = os.getenv("ADMIN_PASSWORD", "1234")

socketio = SocketIO(app, async_mode="threading", cors_allowed_origins="*")

if os.path.exists("/var/data"):
    BASE_PATH = "/var/data"
else:
    BASE_PATH = "data"

UPLOAD_FOLDER = os.path.join(BASE_PATH, "uploads")
DB_FILE = os.path.join(BASE_PATH, "datos.db")

os.makedirs(BASE_PATH, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{DB_FILE}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "connect_args": {
        "timeout": 10
    }
}


@event.listens_for(Engine, "connect")
def configurar_conexion_sqlite(dbapi_connection, connection_record):
    if not isinstance(dbapi_connection, sqlite3.Connection):
        return

    cursor = dbapi_connection.cursor()
    try:
        cursor.execute("PRAGMA busy_timeout = 10000")
    finally:
        cursor.close()

db = SQLAlchemy(app)


TABLAS_PRINCIPALES = {
    "participante",
    "config",
    "handicap",
    "registro_movimiento"
}


class ErrorBackupReset(Exception):
    pass


def es_bloqueo_sqlite(error):
    mensajes = []
    actual = error

    while actual is not None:
        mensajes.append(str(actual).lower())
        siguiente = actual.__cause__ or actual.__context__
        if siguiente is actual:
            break
        actual = siguiente

    texto = " ".join(mensajes)
    return "database is locked" in texto or "database table is locked" in texto


def registrar_error_transaccion(contexto, error):
    if es_bloqueo_sqlite(error):
        app.logger.warning(
            "Bloqueo de SQLite agotó el tiempo de espera en %s",
            contexto
        )
    else:
        app.logger.error(
            "Error importante de base de datos en %s (%s)",
            contexto,
            type(error).__name__
        )


def rollback_transaccion(contexto):
    db.session.rollback()
    app.logger.warning("Rollback ejecutado en %s", contexto)


@app.errorhandler(OperationalError)
def manejar_error_operacional(error):
    rollback_transaccion(f"ruta {request.path}")

    if es_bloqueo_sqlite(error):
        app.logger.warning(
            "Bloqueo de SQLite no resuelto en la ruta %s",
            request.path
        )

        if request.path == "/agregar" or request.path.startswith("/delete/"):
            return jsonify({
                "ok": False,
                "msg": "La base de datos está ocupada. Intentá nuevamente."
            }), 503

        if request.path == "/lista_whatsapp":
            return jsonify({
                "ok": False,
                "msg": "La base de datos está ocupada. Intentá nuevamente."
            }), 503

        return "La base de datos está ocupada. Intentá nuevamente.", 503

    app.logger.error(
        "Error operacional de base de datos en %s (%s)",
        request.path,
        type(error).__name__
    )
    return "No se pudo completar la operación", 503


class Participante(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    apellido = db.Column(db.String(100), nullable=False)
    dni_matricula = db.Column(db.String(50))
    matricula = db.Column(db.String(50))
    nombre_normalizado = db.Column(db.String(100))
    apellido_normalizado = db.Column(db.String(100))
    dni_matricula_normalizado = db.Column(db.String(50))
    matricula_normalizada = db.Column(db.String(50))
    asistencia = db.Column(db.String(100))
    equipo = db.Column(db.String(50))


class Config(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    titulo = db.Column(db.String(200), default="🏌️ Torneo Matungo")
    subtitulo = db.Column(db.String(200), default="")
    subtitulo2 = db.Column(db.String(200), default="")
    subtitulo3 = db.Column(db.String(200), default="")
    opciones_menu = db.Column(db.Text, default="8:00 AM,9:00 AM,10:00 AM")
    menu_activo = db.Column(db.Boolean, default=True)
    cierre_inscripcion = db.Column(db.String(30), default="")
    whatsapp_activo = db.Column(db.Boolean, default=True)


class Handicap(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100))
    hdcp = db.Column(db.String(10))

class RegistroMovimiento(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    participante_id = db.Column(db.Integer)
    accion = db.Column(db.String(30), nullable=False)
    nombre = db.Column(db.String(100))
    apellido = db.Column(db.String(100))
    dni_matricula = db.Column(db.String(50))
    matricula = db.Column(db.String(50))
    equipo = db.Column(db.String(50))
    asistencia = db.Column(db.String(100))
    detalle = db.Column(db.String(300))
    fecha_hora = db.Column(db.String(30), nullable=False)

def normalizar_nombre(t):
    return " ".join((t or "").strip().split()).casefold()


def normalizar_dni_matricula(t):
    texto = (t or "").strip()

    if not texto:
        return None

    if re.fullmatch(r"[\d\s.\-]+", texto):
        return re.sub(r"[\s.\-]", "", texto)

    return texto.casefold()


def normalizar_matricula(t):
    texto = (t or "").strip()
    return texto.casefold() if texto else None


def enmascarar_dni_matricula_normalizado(valor):
    if not valor:
        return "(vacío)"

    visibles = valor[-4:]
    return "*" * max(0, len(valor) - len(visibles)) + visibles


def migrar_columna(conn, tabla, columna, definicion):
    columnas = [row[1] for row in conn.exec_driver_sql(f"PRAGMA table_info({tabla})")]
    if columna not in columnas:
        conn.exec_driver_sql(f"ALTER TABLE {tabla} ADD COLUMN {columna} {definicion}")


def completar_normalizaciones_participantes(conn):
    participantes = conn.exec_driver_sql("""
        SELECT id, nombre, apellido, dni_matricula, matricula,
               nombre_normalizado, apellido_normalizado,
               dni_matricula_normalizado, matricula_normalizada
        FROM participante
    """).fetchall()

    for participante in participantes:
        valores_normalizados = {
            "id": participante[0],
            "nombre_normalizado": normalizar_nombre(participante[1]),
            "apellido_normalizado": normalizar_nombre(participante[2]),
            "dni_matricula_normalizado": normalizar_dni_matricula(
                participante[3]
            ),
            "matricula_normalizada": normalizar_matricula(participante[4])
        }

        if (
            participante[5] != valores_normalizados["nombre_normalizado"]
            or participante[6] != valores_normalizados["apellido_normalizado"]
            or participante[7]
            != valores_normalizados["dni_matricula_normalizado"]
            or participante[8] != valores_normalizados["matricula_normalizada"]
        ):
            conn.exec_driver_sql("""
                UPDATE participante
                SET nombre_normalizado = :nombre_normalizado,
                    apellido_normalizado = :apellido_normalizado,
                    dni_matricula_normalizado = :dni_matricula_normalizado,
                    matricula_normalizada = :matricula_normalizada
                WHERE id = :id
            """, valores_normalizados)


def buscar_duplicados_normalizados(conn):
    filas = conn.exec_driver_sql("""
        SELECT id, nombre_normalizado, apellido_normalizado,
               dni_matricula_normalizado, matricula_normalizada
        FROM participante
        ORDER BY id
    """).fetchall()

    nombres = {}
    documentos = {}
    matriculas = {}

    for fila in filas:
        clave_nombre = (fila[1], fila[2])
        nombres.setdefault(clave_nombre, []).append(fila[0])

        if fila[3]:
            documentos.setdefault(fila[3], []).append(fila[0])

        if fila[4]:
            matriculas.setdefault(fila[4], []).append(fila[0])

    duplicados_nombres = [
        ids for ids in nombres.values() if len(ids) > 1
    ]
    duplicados_documentos = [
        {
            "ids": ids,
            "dni_matricula": documento
        }
        for documento, ids in documentos.items()
        if len(ids) > 1
    ]
    duplicados_matriculas = [
        {
            "ids": ids,
            "matricula": matricula
        }
        for matricula, ids in matriculas.items()
        if len(ids) > 1
    ]

    return duplicados_nombres, duplicados_documentos, duplicados_matriculas


def crear_triggers_unicidad_participantes(conn):
    conn.exec_driver_sql("""
        CREATE TRIGGER IF NOT EXISTS trg_participante_nombre_unico_insert
        BEFORE INSERT ON participante
        WHEN EXISTS (
            SELECT 1 FROM participante
            WHERE nombre_normalizado = NEW.nombre_normalizado
              AND apellido_normalizado = NEW.apellido_normalizado
        )
        BEGIN
            SELECT RAISE(ABORT, 'nombre_normalizado_duplicado');
        END
    """)
    conn.exec_driver_sql("""
        CREATE TRIGGER IF NOT EXISTS trg_participante_nombre_unico_update
        BEFORE UPDATE OF nombre_normalizado, apellido_normalizado ON participante
        WHEN EXISTS (
            SELECT 1 FROM participante
            WHERE nombre_normalizado = NEW.nombre_normalizado
              AND apellido_normalizado = NEW.apellido_normalizado
              AND id != NEW.id
        )
        BEGIN
            SELECT RAISE(ABORT, 'nombre_normalizado_duplicado');
        END
    """)
    conn.exec_driver_sql("""
        CREATE TRIGGER IF NOT EXISTS trg_participante_dni_unico_insert
        BEFORE INSERT ON participante
        WHEN EXISTS (
            SELECT 1 FROM participante
            WHERE dni_matricula_normalizado = NEW.dni_matricula_normalizado
        )
        BEGIN
            SELECT RAISE(ABORT, 'dni_matricula_normalizado_duplicado');
        END
    """)
    conn.exec_driver_sql("""
        CREATE TRIGGER IF NOT EXISTS trg_participante_dni_unico_update
        BEFORE UPDATE OF dni_matricula_normalizado ON participante
        WHEN EXISTS (
            SELECT 1 FROM participante
            WHERE dni_matricula_normalizado = NEW.dni_matricula_normalizado
              AND id != NEW.id
        )
        BEGIN
            SELECT RAISE(ABORT, 'dni_matricula_normalizado_duplicado');
        END
    """)
    conn.exec_driver_sql("""
        CREATE TRIGGER IF NOT EXISTS trg_participante_matricula_unica_insert
        BEFORE INSERT ON participante
        WHEN NEW.matricula_normalizada IS NOT NULL
         AND EXISTS (
            SELECT 1 FROM participante
            WHERE matricula_normalizada = NEW.matricula_normalizada
        )
        BEGIN
            SELECT RAISE(ABORT, 'matricula_normalizada_duplicada');
        END
    """)
    conn.exec_driver_sql("""
        CREATE TRIGGER IF NOT EXISTS trg_participante_matricula_unica_update
        BEFORE UPDATE OF matricula_normalizada ON participante
        WHEN NEW.matricula_normalizada IS NOT NULL
         AND EXISTS (
            SELECT 1 FROM participante
            WHERE matricula_normalizada = NEW.matricula_normalizada
              AND id != NEW.id
        )
        BEGIN
            SELECT RAISE(ABORT, 'matricula_normalizada_duplicada');
        END
    """)


def crear_indices_unicos_si_es_posible(conn):
    duplicados_nombres, duplicados_documentos, duplicados_matriculas = (
        buscar_duplicados_normalizados(conn)
    )

    if duplicados_nombres:
        app.logger.error(
            "No se creó el índice único de nombres normalizados. "
            "Registros a revisar manualmente: %s",
            duplicados_nombres
        )
        indice_nombres_creado = False
    else:
        conn.exec_driver_sql("""
            CREATE UNIQUE INDEX IF NOT EXISTS uq_participante_nombre_normalizado
            ON participante (nombre_normalizado, apellido_normalizado)
        """)
        indice_nombres_creado = True

    if duplicados_documentos:
        app.logger.error(
            "No se creó el índice único de DNI o matrícula. "
            "Registros a revisar manualmente: %s",
            [
                {
                    "ids": duplicado["ids"],
                    "valor_enmascarado": (
                        enmascarar_dni_matricula_normalizado(
                            duplicado["dni_matricula"]
                        )
                    )
                }
                for duplicado in duplicados_documentos
            ]
        )
        indice_documentos_creado = False
    else:
        conn.exec_driver_sql("""
            CREATE UNIQUE INDEX IF NOT EXISTS uq_participante_dni_normalizado
            ON participante (dni_matricula_normalizado)
        """)
        indice_documentos_creado = True

    if duplicados_matriculas:
        app.logger.error(
            "No se creó el índice único de matrículas normalizadas. "
            "Registros a revisar manualmente: %s",
            [duplicado["ids"] for duplicado in duplicados_matriculas]
        )
        indice_matriculas_creado = False
    else:
        conn.exec_driver_sql("""
            CREATE UNIQUE INDEX IF NOT EXISTS uq_participante_matricula_normalizada
            ON participante (matricula_normalizada)
            WHERE matricula_normalizada IS NOT NULL
        """)
        indice_matriculas_creado = True

    crear_triggers_unicidad_participantes(conn)

    configuraciones = conn.exec_driver_sql(
        "SELECT id FROM config ORDER BY id"
    ).fetchall()

    if len(configuraciones) > 1:
        app.logger.error(
            "No se creo el indice singleton de configuracion: existen %s filas",
            len(configuraciones)
        )
    else:
        conn.exec_driver_sql("""
            CREATE UNIQUE INDEX IF NOT EXISTS uq_config_singleton
            ON config ((1))
        """)

    return (
        indice_nombres_creado,
        indice_documentos_creado,
        indice_matriculas_creado
    )


with app.app_context():
    db.create_all()

    with db.engine.begin() as conn:
        migrar_columna(conn, "config", "subtitulo2", "VARCHAR(200) DEFAULT ''")
        migrar_columna(conn, "config", "subtitulo3", "VARCHAR(200) DEFAULT ''")
        migrar_columna(conn, "config", "opciones_menu", "TEXT DEFAULT '8:00 AM,9:00 AM,10:00 AM'")
        migrar_columna(conn, "config", "menu_activo", "BOOLEAN DEFAULT 1")
        migrar_columna(conn, "config", "whatsapp_activo", "BOOLEAN DEFAULT 1")
        migrar_columna(conn, "config", "cierre_inscripcion", "VARCHAR(30) DEFAULT ''")

        migrar_columna(conn, "participante", "asistencia", "VARCHAR(100)")
        migrar_columna(conn, "participante", "equipo", "VARCHAR(50)")
        migrar_columna(conn, "participante", "dni_matricula", "VARCHAR(50)")
        migrar_columna(conn, "participante", "matricula", "VARCHAR(50)")
        migrar_columna(
            conn,
            "participante",
            "nombre_normalizado",
            "VARCHAR(100)"
        )
        migrar_columna(
            conn,
            "participante",
            "apellido_normalizado",
            "VARCHAR(100)"
        )
        migrar_columna(
            conn,
            "participante",
            "dni_matricula_normalizado",
            "VARCHAR(50)"
        )
        migrar_columna(
            conn,
            "participante",
            "matricula_normalizada",
            "VARCHAR(50)"
        )
        migrar_columna(
            conn,
            "registro_movimiento",
            "matricula",
            "VARCHAR(50)"
        )

        completar_normalizaciones_participantes(conn)
        indices_unicos = crear_indices_unicos_si_es_posible(conn)
        app.config["PARTICIPANTE_NAME_UNIQUE_INDEX_READY"] = (
            indices_unicos[0]
        )
        app.config["PARTICIPANTE_DNI_UNIQUE_INDEX_READY"] = (
            indices_unicos[1]
        )
        app.config["PARTICIPANTE_MATRICULA_UNIQUE_INDEX_READY"] = (
            indices_unicos[2]
        )

    if not Config.query.first():
        try:
            db.session.add(Config())
            db.session.commit()
        except IntegrityError:
            rollback_transaccion("creacion concurrente de configuracion inicial")
        except Exception as error:
            rollback_transaccion("creacion de configuracion inicial")
            registrar_error_transaccion("creacion de configuracion inicial", error)
            raise


def validar_dni(t):
    return re.fullmatch(r"\d{1,8}", t or "")


def mensaje_error_integridad_participante(error):
    detalle = str(getattr(error, "orig", error)).lower()

    if "dni_matricula_normalizado" in detalle:
        return "El DNI ingresado ya se encuentra registrado."

    if "matricula_normalizada" in detalle:
        return "La matrícula ingresada ya se encuentra registrada."

    if (
        "nombre_normalizado" in detalle
        or "uq_participante_nombre" in detalle
        or "participante.nombre, participante.apellido" in detalle
    ):
        return "Ya está anotado"

    return "No se pudo completar la inscripción. Intentá nuevamente."


def solo_letras(t):
    return re.match(r"^[A-Za-zÁÉÍÓÚáéíóúÑñ ]+$", t or "")


def participantes_ordenados():
    return Participante.query.order_by(
        Participante.apellido.asc(),
        Participante.nombre.asc()
    ).all()


def participantes_por_equipo(participantes):
    return {
        "team22": [
            p for p in participantes
            if p.equipo == "Team 22"
        ],
        "aguilas": [
            p for p in participantes
            if p.equipo == "Águilas"
        ],
        "invitados": [
            p for p in participantes
            if p.equipo == "Invitado"
        ]
    }


def agregar_equipo_lista_whatsapp(lineas, titulo, participantes):
    lineas.append(f"{titulo} ({len(participantes)})")
    lineas.append("")

    if participantes:
        for participante in participantes:
            nombre_completo = (
                f"{participante.nombre} {participante.apellido}"
            ).strip()
            lineas.append(f"• {nombre_completo}")
    else:
        lineas.append("Sin inscriptos")

    lineas.append("")
    lineas.append("━━━━━━━━━━━━━━━━━━")
    lineas.append("")


def generar_texto_whatsapp(config, participantes):
    equipos = participantes_por_equipo(participantes)

    titulo = (
        config.titulo.strip()
        if config and config.titulo
        else "Matungo Golf Tour"
    )

    lineas = [
        f"🏌️ *{titulo}*",
        "",
        "📅 *Inscriptos actualizados*",
        f"🕒 {ahora_argentina().strftime('%d/%m/%Y - %H:%M hs')}",
        "",
        "━━━━━━━━━━━━━━━━━━",
        ""
    ]

    agregar_equipo_lista_whatsapp(
        lineas,
        "🔵 *TEAM 22*",
        equipos["team22"]
    )

    agregar_equipo_lista_whatsapp(
        lineas,
        "🦅 *ÁGUILAS*",
        equipos["aguilas"]
    )

    agregar_equipo_lista_whatsapp(
        lineas,
        "👤 *INVITADOS*",
        equipos["invitados"]
    )

    lineas.extend([
        f"👥 *TOTAL INSCRIPTOS:* {len(participantes)}",
        "",
        (
            "⚠️ Si realizaste la inscripción y no figurás en este listado, "
            "comunicate inmediatamente con la organización."
        )
    ])

    return "\n".join(lineas)


def ahora_argentina():
    return datetime.utcnow() - timedelta(hours=3)


def obtener_ruta_base_sqlite():
    bases = db.session.connection().exec_driver_sql(
        "PRAGMA database_list"
    ).fetchall()

    for base in bases:
        if base[1] == "main" and base[2]:
            return Path(base[2]).resolve()

    raise ErrorBackupReset("No se pudo determinar la base SQLite activa")


def crear_nombre_backup_reset(carpeta):
    momento = ahora_argentina()
    nombre = f"datos_antes_reset_{momento.strftime('%Y-%m-%d_%H-%M-%S')}.db"
    destino = carpeta / nombre

    if not destino.exists():
        return destino

    nombre = (
        "datos_antes_reset_"
        f"{momento.strftime('%Y-%m-%d_%H-%M-%S_%f')}.db"
    )
    destino = carpeta / nombre
    contador = 1

    while destino.exists():
        destino = carpeta / f"{Path(nombre).stem}_{contador}.db"
        contador += 1

    return destino


def verificar_backup_reset(ruta_backup, cantidad_participantes):
    if not ruta_backup.exists() or ruta_backup.stat().st_size <= 0:
        raise ErrorBackupReset("El archivo de backup no existe o está vacío")

    conexion = None

    try:
        conexion = sqlite3.connect(
            f"file:{ruta_backup.as_posix()}?mode=ro",
            uri=True,
            timeout=10
        )
        integridad = conexion.execute("PRAGMA integrity_check").fetchone()

        if not integridad or integridad[0].lower() != "ok":
            raise ErrorBackupReset("La verificación de integridad falló")

        tablas = {
            fila[0]
            for fila in conexion.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            )
        }

        if not TABLAS_PRINCIPALES.issubset(tablas):
            raise ErrorBackupReset("Faltan tablas principales en el backup")

        cantidad_backup = conexion.execute(
            "SELECT COUNT(*) FROM participante"
        ).fetchone()[0]

        if cantidad_backup != cantidad_participantes:
            raise ErrorBackupReset(
                "La cantidad de participantes del backup no coincide"
            )
    except sqlite3.Error as error:
        raise ErrorBackupReset("No se pudo validar la base de backup") from error
    finally:
        if conexion is not None:
            conexion.close()


def crear_backup_reset(cantidad_participantes):
    ruta_base = obtener_ruta_base_sqlite()
    carpeta_backups = ruta_base.parent / "backups_reset"
    carpeta_backups.mkdir(parents=True, exist_ok=True)
    ruta_backup = crear_nombre_backup_reset(carpeta_backups)

    origen = None
    destino = None
    error_creacion = None

    try:
        origen = sqlite3.connect(
            f"file:{ruta_base.as_posix()}?mode=ro",
            uri=True,
            timeout=10
        )
        destino = sqlite3.connect(str(ruta_backup), timeout=10)
        origen.backup(destino)
    except Exception as error:
        error_creacion = error
    finally:
        if destino is not None:
            destino.close()
        if origen is not None:
            origen.close()

    if error_creacion is not None:
        try:
            if ruta_backup.exists():
                ruta_backup.unlink()
        except OSError as error:
            app.logger.error(
                "No se pudo eliminar un backup incompleto (%s)",
                type(error).__name__
            )
        raise ErrorBackupReset(
            "No se pudo crear el backup SQLite"
        ) from error_creacion

    try:
        verificar_backup_reset(ruta_backup, cantidad_participantes)
    except Exception:
        try:
            if ruta_backup.exists():
                ruta_backup.unlink()
        except OSError as error:
            app.logger.error(
                "No se pudo eliminar un backup incompleto (%s)",
                type(error).__name__
            )
        raise

    return ruta_backup

def registrar_movimiento(
    accion,
    participante=None,
    detalle=""
):
    registro = RegistroMovimiento(
        participante_id=participante.id if participante else None,
        accion=accion,
        nombre=participante.nombre if participante else "",
        apellido=participante.apellido if participante else "",
        dni_matricula=participante.dni_matricula if participante else "",
        matricula=participante.matricula if participante else "",
        equipo=participante.equipo if participante else "",
        asistencia=participante.asistencia if participante else "",
        detalle=detalle,
        fecha_hora=ahora_argentina().strftime("%d/%m/%Y %H:%M:%S")
    )

    db.session.add(registro)

def inscripcion_cerrada(config):
    if not config.cierre_inscripcion:
        return False

    try:
        cierre = datetime.fromisoformat(config.cierre_inscripcion)
        return ahora_argentina() >= cierre
    except Exception:
        return False


@app.route("/")
def index():
    config = Config.query.first()

    menu_opciones = []
    if config and config.opciones_menu:
        menu_opciones = [x.strip() for x in config.opciones_menu.split(",") if x.strip()]

    bg_path = "/static_bg" if os.path.exists(os.path.join(UPLOAD_FOLDER, "fondo.jpg")) else None

    return render_template(
        "index.html",
        participantes=participantes_ordenados(),
        handicaps=Handicap.query.order_by(Handicap.nombre.asc()).all(),
        registros=RegistroMovimiento.query.order_by(
            RegistroMovimiento.id.desc()
        ).limit(200).all(),
        menu_opciones=menu_opciones,
        menu_activo=config.menu_activo,
        admin=session.get("admin", False),
        bg_path=bg_path,
        inscripcion_cerrada=inscripcion_cerrada(config),
        config=config
    )


@app.route("/participantes")
def participantes():
    data = participantes_ordenados()
    return jsonify([
        {
            "id": p.id,
            "nombre": p.nombre,
            "apellido": p.apellido,
            "equipo": p.equipo or ""
        }
        for p in data
    ])


@app.route("/lista_whatsapp")
def lista_whatsapp():
    if not session.get("admin"):
        return jsonify({
            "ok": False,
            "msg": "No autorizado"
        }), 403

    config = Config.query.first()
    participantes = participantes_ordenados()
    texto = generar_texto_whatsapp(config, participantes)

    return jsonify({
        "ok": True,
        "texto": texto,
        "total": len(participantes),
        "generado": ahora_argentina().isoformat(timespec="seconds")
    })


@app.route("/registros")
def registros():
    if not session.get("admin"):
        return jsonify([])

    data = RegistroMovimiento.query.order_by(
        RegistroMovimiento.id.desc()
    ).limit(200).all()

    return jsonify([
        {
            "id": r.id,
            "accion": r.accion or "",
            "fecha_hora": r.fecha_hora or "",
            "nombre": r.nombre or "",
            "apellido": r.apellido or "",
            "dni_matricula": r.dni_matricula or "",
            "matricula": r.matricula or "",
            "equipo": r.equipo or "",
            "asistencia": r.asistencia or "",
            "detalle": r.detalle or ""
        }
        for r in data
    ])

@app.route("/agregar", methods=["POST"])
def agregar():
    config = Config.query.first()

    nombre = request.form.get("nombre", "").strip()
    apellido = request.form.get("apellido", "").strip()
    dni_matricula = request.form.get("dni_matricula", "").strip()
    matricula = request.form.get("matricula", "").strip()
    asistencia = request.form.get("asistencia", "").strip()
    equipo = request.form.get("equipo", "").strip()
    nombre_normalizado = normalizar_nombre(nombre)
    apellido_normalizado = normalizar_nombre(apellido)
    dni_matricula_normalizado = normalizar_dni_matricula(dni_matricula)
    matricula_normalizada = normalizar_matricula(matricula)

    if inscripcion_cerrada(config):
        return jsonify({"ok": False, "msg": "La inscripción ya está cerrada"}), 400

    if not equipo:
        return jsonify({"ok": False, "msg": "Falta elegir equipo"}), 400

    if config.menu_activo and not asistencia:
        return jsonify({"ok": False, "msg": "Falta elegir menú"}), 400

    if not solo_letras(nombre) or not solo_letras(apellido):
        return jsonify({"ok": False, "msg": "Nombre o apellido inválido"}), 400

    if not dni_matricula:
        return jsonify({"ok": False, "msg": "Debe ingresar DNI"}), 400

    if not validar_dni(dni_matricula_normalizado):
        return jsonify({"ok": False, "msg": "El DNI debe contener solo números y hasta 8 dígitos"}), 400

    if Participante.query.filter_by(
        nombre_normalizado=nombre_normalizado,
        apellido_normalizado=apellido_normalizado
    ).first():
        return jsonify({"ok": False, "msg": "Ya está anotado"}), 400

    if Participante.query.filter_by(
        dni_matricula_normalizado=dni_matricula_normalizado
    ).first():
        return jsonify({
            "ok": False,
            "msg": "El DNI ingresado ya se encuentra registrado."
        }), 400

    if matricula_normalizada and Participante.query.filter_by(
        matricula_normalizada=matricula_normalizada
    ).first():
        return jsonify({
            "ok": False,
            "msg": "La matrícula ingresada ya se encuentra registrada."
        }), 400

    nuevo = Participante(
        nombre=nombre,
        apellido=apellido,
        dni_matricula=dni_matricula,
        matricula=matricula or None,
        nombre_normalizado=nombre_normalizado,
        apellido_normalizado=apellido_normalizado,
        dni_matricula_normalizado=dni_matricula_normalizado,
        matricula_normalizada=matricula_normalizada,
        asistencia=asistencia,
        equipo=equipo
    )

    try:
        db.session.add(nuevo)
        db.session.flush()

        registrar_movimiento(
            accion="ALTA",
            participante=nuevo,
            detalle="Inscripción completada correctamente"
        )

        db.session.commit()
    except IntegrityError as error:
        rollback_transaccion("alta duplicada de participante")
        return jsonify({
            "ok": False,
            "msg": mensaje_error_integridad_participante(error)
        }), 409
    except Exception as error:
        rollback_transaccion("registro de participante")
        registrar_error_transaccion("registro de participante", error)
        return jsonify({
            "ok": False,
            "msg": "No se pudo completar la inscripción. Intentá nuevamente."
        }), 503

    socketio.emit("actualizar_lista")
    socketio.emit("actualizar_registro")

    return jsonify({
        "ok": True,
        "nombre_completo": f"{nombre} {apellido}",
        "equipo": equipo,
        "menu": asistencia,
        "mensaje_evento": config.subtitulo2 or config.titulo
    })


@app.route("/admin", methods=["POST"])
def admin_login():
    if request.form.get("password") == ADMIN_PASSWORD:
        session["admin"] = True
    return redirect("/")


@app.route("/admin-secret")
def admin_secret():
    return render_template("admin_login.html")


@app.route("/logout")
def logout():
    session.pop("admin", None)
    return redirect("/")


@app.route("/update_config", methods=["POST"])
def update_config():
    if not session.get("admin"):
        return "No autorizado", 403

    config = Config.query.first()

    try:
        config.titulo = request.form.get("titulo", "")
        config.subtitulo = request.form.get("subtitulo", "")
        config.subtitulo2 = request.form.get("subtitulo2", "")
        config.subtitulo3 = request.form.get("subtitulo3", "")
        config.cierre_inscripcion = request.form.get("cierre_inscripcion", "").strip()
        config.opciones_menu = request.form.get("opciones_menu", "")
        config.menu_activo = bool(request.form.get("menu_activo"))
        config.whatsapp_activo = bool(request.form.get("whatsapp_activo"))

        db.session.commit()
    except Exception as error:
        rollback_transaccion("actualizacion de configuracion")
        registrar_error_transaccion("actualizacion de configuracion", error)
        return "No se pudo guardar la configuración", 503
    app.logger.info("Configuracion actualizada por administrador")
    socketio.emit("actualizar_lista")

    return redirect("/")


@app.route("/delete/<int:id>", methods=["POST"])
def delete(id):
    if not session.get("admin"):
        return jsonify({
            "ok": False,
            "msg": "No autorizado"
        }), 403

    participante = db.session.get(Participante, id)

    if not participante:
        return jsonify({
            "ok": False,
            "msg": "Jugador no encontrado"
        }), 404

    try:
        registrar_movimiento(
            accion="ELIMINADO",
            participante=participante,
            detalle="Eliminado desde el panel administrador"
        )

        db.session.delete(participante)
        db.session.commit()
    except Exception as error:
        rollback_transaccion("eliminacion de participante")
        registrar_error_transaccion("eliminacion de participante", error)
        return jsonify({
            "ok": False,
            "msg": "No se pudo eliminar el jugador"
        }), 503

    app.logger.info(
        "Operacion administrativa de eliminacion completada: registros=%s",
        1
    )

    socketio.emit("actualizar_lista")
    socketio.emit("actualizar_registro")

    return jsonify({"ok": True})


@app.route("/reset")
def reset():
    if not session.get("admin"):
        return "No autorizado", 403

    ruta_backup = None

    try:
        db.session.execute(db.text("BEGIN IMMEDIATE"))
        participantes = Participante.query.all()
        cantidad = len(participantes)

        app.logger.info(
            "Inicio de backup previo al reset: participantes=%s",
            cantidad
        )

        try:
            ruta_backup = crear_backup_reset(cantidad)
            app.logger.info(
                "Backup previo al reset verificado: archivo=%s participantes=%s",
                ruta_backup.name,
                cantidad
            )
        except Exception as error:
            rollback_transaccion("backup previo al reset")
            app.logger.error(
                "Error importante en backup previo al reset (%s)",
                type(error).__name__
            )
            flash(
                "No se pudo crear o verificar el backup. "
                "La lista no fue reseteada."
            )
            return redirect("/")

        for participante in participantes:
            registrar_movimiento(
                accion="RESET",
                participante=participante,
                detalle="Eliminado durante un reset general"
            )

        Participante.query.delete(synchronize_session=False)

        registro_resumen = RegistroMovimiento(
            participante_id=None,
            accion="RESET GENERAL",
            nombre="",
            apellido="",
            dni_matricula="",
            matricula="",
            equipo="",
            asistencia="",
            detalle=f"Se eliminaron {cantidad} participantes",
            fecha_hora=ahora_argentina().strftime("%d/%m/%Y %H:%M:%S")
        )

        db.session.add(registro_resumen)
        db.session.commit()
    except Exception as error:
        rollback_transaccion("reset general")
        registrar_error_transaccion("reset general", error)

        if ruta_backup is not None:
            flash(
                "El backup fue creado correctamente, pero la lista no pudo "
                f"resetearse. Backup creado: {ruta_backup.name}"
            )
        else:
            flash(
                "No se pudo crear o verificar el backup. "
                "La lista no fue reseteada."
            )
        return redirect("/")

    app.logger.info(
        "Reset administrativo completado: registros_eliminados=%s",
        cantidad
    )
    flash(
        "La lista fue reseteada correctamente. "
        f"Se creó un backup automático con {cantidad} participantes. "
        f"Backup creado: {ruta_backup.name}"
    )

    socketio.emit("actualizar_lista")
    socketio.emit("actualizar_registro")

    return redirect("/")

@app.route("/export")
def export():
    if not session.get("admin"):
        return "No autorizado", 403

    try:
        respuesta, cantidad = crear_exportacion_participantes()
        app.logger.info(
            "Exportacion administrativa de participantes completada: registros=%s",
            cantidad
        )
        return respuesta
    except Exception as error:
        app.logger.error(
            "Error de exportacion de participantes (%s)",
            type(error).__name__
        )
        return "No se pudo generar la exportación", 500


def crear_exportacion_participantes():
    wb = Workbook()
    ws_default = wb.active
    wb.remove(ws_default)

    def crear_hoja(nombre_hoja, participantes):
        ws = wb.create_sheet(title=nombre_hoja)
        ws.append([
            "Nombre", "Apellido", "DNI", "Matrícula", "Equipo", "Menú"
        ])

        for p in participantes:
            ws.append([
                p.nombre,
                p.apellido,
                p.dni_matricula,
                p.matricula,
                p.equipo,
                p.asistencia
            ])

    todos = participantes_ordenados()
    team22 = [p for p in todos if p.equipo == "Team 22"]
    aguilas = [p for p in todos if p.equipo == "Águilas"]
    invitados = [p for p in todos if p.equipo == "Invitado"]

    crear_hoja("General", todos)
    crear_hoja("Team 22", team22)
    crear_hoja("Aguilas", aguilas)
    crear_hoja("Invitados", invitados)

    respuesta = enviar_workbook(wb, "participantes.xlsx")
    return respuesta, len(todos)

@app.route("/export_historial")
def export_historial():
    if not session.get("admin"):
        return "No autorizado", 403

    try:
        respuesta, cantidad = crear_exportacion_historial()
        app.logger.info(
            "Exportacion administrativa de historial completada: registros=%s",
            cantidad
        )
        return respuesta
    except Exception as error:
        app.logger.error(
            "Error de exportacion de historial (%s)",
            type(error).__name__
        )
        return "No se pudo generar la exportación", 500


def crear_exportacion_historial():
    registros = RegistroMovimiento.query.order_by(
        RegistroMovimiento.id.asc()
    ).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Historial"

    ws.append([
        "Fecha y hora",
        "Acción",
        "Nombre",
        "Apellido",
        "DNI",
        "Matrícula",
        "Equipo",
        "Menú",
        "Detalle"
    ])

    for r in registros:
        ws.append([
            r.fecha_hora,
            r.accion,
            r.nombre,
            r.apellido,
            r.dni_matricula,
            r.matricula,
            r.equipo,
            r.asistencia,
            r.detalle
        ])

    respuesta = enviar_workbook(wb, "historial_movimientos.xlsx")
    return respuesta, len(registros)


def enviar_workbook(wb, nombre_descarga):
    archivo = BytesIO()

    try:
        wb.save(archivo)
    except Exception:
        archivo.close()
        raise
    finally:
        wb.close()

    archivo.seek(0)
    return send_file(
        archivo,
        as_attachment=True,
        download_name=nombre_descarga,
        mimetype=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )


@app.route("/upload_hdcp", methods=["POST"])
def upload_hdcp():
    if not session.get("admin"):
        return "No autorizado", 403

    file = request.files.get("file")
    if not file:
        return redirect("/")

    wb = None
    cantidad_importados = 0

    try:
        wb = load_workbook(file)
        ws = wb.active

        cantidad_reemplazados = Handicap.query.delete(synchronize_session=False)

        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue

            if not row or len(row) < 2:
                continue

            nombre, hdcp = row[0], row[1]

            if nombre and hdcp:
                db.session.add(Handicap(
                    nombre=str(nombre).strip(),
                    hdcp=str(hdcp).strip()
                ))
                cantidad_importados += 1

        db.session.commit()
        app.logger.info(
            "Importacion administrativa de handicaps completada: "
            "registros_importados=%s registros_reemplazados=%s",
            cantidad_importados,
            cantidad_reemplazados
        )
    except OperationalError as error:
        rollback_transaccion("importacion de handicaps")
        registrar_error_transaccion("importacion de handicaps", error)
        return "La base de datos está ocupada. Intentá nuevamente.", 503
    except Exception as error:
        rollback_transaccion("importacion de handicaps")
        app.logger.error(
            "Error importante durante la importacion de handicaps (%s)",
            type(error).__name__
        )
        return "No se pudo importar el archivo de handicap", 400
    finally:
        if wb is not None:
            wb.close()

    return redirect("/")


@app.route("/upload_bg", methods=["POST"])
def upload_bg():
    if not session.get("admin"):
        return "No autorizado", 403

    file = request.files.get("imagen")

    if file and file.filename.lower().endswith((".png", ".jpg", ".jpeg")):
        try:
            file.save(os.path.join(UPLOAD_FOLDER, "fondo.jpg"))
            app.logger.info("Fondo actualizado por administrador")
        except Exception as error:
            app.logger.error(
                "Error importante al actualizar el fondo (%s)",
                type(error).__name__
            )
            return "No se pudo actualizar el fondo", 500

    return redirect("/")


@app.route("/static_bg")
def bg():
    return send_file(os.path.join(UPLOAD_FOLDER, "fondo.jpg"))


if __name__ == "__main__":
    socketio.run(
        app,
        host="0.0.0.0",
        port=int(os.getenv("PORT", 10000)),
        debug=True,
        allow_unsafe_werkzeug=True
    )
